# %%
from sqlalchemy import create_engine
import pandas as pd
import numpy as np
from datetime import date, datetime, timedelta
from datetime import datetime
from dcon import get_data
from mail import send_mail
import time
import os
import glob
# import holiday file from E\Config
import sys
sys.path.append(r'E:\_Config')  # Assuming 'holiday.py' is in this directory
from holiday import holiday


# check if today is friday or holiday. if holiday or friday then exit the program
def is_today_holiday(holiday_list):
    today_date = datetime.now().date().strftime('%Y-%m-%d')
    return today_date in holiday_list

def is_today_friday():
    return datetime.now().weekday() == 4  # Friday is 4

if is_today_holiday(holiday()):
    print("Today is a holiday. Exiting program.")
    sys.exit()

if is_today_friday():
    print("Today is Friday. Exiting program.")
    sys.exit()



zid = 100001
# district = ('Sylhet Retail', 'Rahima enterprise', 'District', 'Chittagong retail', 'Imamgonj') 
salesman_info_dict = {
'XXXXXX':'XXXXXX'
}

salesman = tuple(salesman_info_dict.keys())

# Get Today Date
now_date = datetime.today()
today_date = now_date.strftime("%Y-%m-%d")
# today_date = '2023-12-05'
print (today_date)




# %%
twenty_days_ago = now_date - timedelta (20)
twenty_days_ago = twenty_days_ago.strftime("%Y-%m-%d")


# %%
#%%
to_salesman_query = f"""
                        SELECT
                        opdor.xdornum AS do_number,
                        opdor.xcus,
                        CONCAT(opdor.xcus, '--', cacus.xshort) AS customer,
                        cacus.xadd1 AS address,
                        cacus.xtaxnum AS mobile_number,
                        cacus.xcity AS city,
                        opdor.xsp,
                        prmst.xname AS SName,
                        opdor.xtotamt AS total_DO_amt,
                        opdor.xdate AS goods_receive_date
                        FROM
                        opdor
                        LEFT JOIN cacus ON opdor.xcus = cacus.xcus AND opdor.zid = cacus.zid
                        JOIN prmst ON opdor.xsp = prmst.xemp
                        WHERE
                        opdor.zid = {zid}
                        AND cacus.zid = {zid}
                        AND prmst.zid = {zid}  -- Corrected the logical error here
                        AND opdor.xdate >= '{twenty_days_ago}'
        """

df_send_to_salesman = get_data (to_salesman_query , fetchOneOrAll = 'all' , df = True, conn = "local")

df_send_to_salesman

# %%
df_send_to_salesman['goods_receive_date'] = pd.to_datetime(df_send_to_salesman['goods_receive_date']) + pd.Timedelta(days=3)
df_send_to_salesman



# %%

df_send_to_salesman['today_date'] = today_date
df_send_to_salesman['date_diff'] = pd.to_datetime(df_send_to_salesman['today_date']) - pd.to_datetime (df_send_to_salesman['goods_receive_date'])
df_send_to_salesman = df_send_to_salesman.sort_values(by=['customer', 'date_diff'], ascending = False)

# this line will need if criteria given
df_send_to_salesman = df_send_to_salesman[df_send_to_salesman['date_diff'].dt.days <= 20].sort_values(by=['customer', 'date_diff']).drop(columns=['today_date'])
df_send_to_salesman


# %%
get_customer_id = tuple(df_send_to_salesman['xcus'].to_list())


# now get their untill previous days balance
today_balance_col = f"till_{today_date}_balance".replace("-","_")

balance_query = f"""
    SELECT gldetail.xsub AS xcus, SUM(gldetail.xprime) AS balance_till_today
    FROM glheader
    JOIN gldetail ON glheader.xvoucher = gldetail.xvoucher
    JOIN cacus ON gldetail.xsub = cacus.xcus
    WHERE glheader.zid = {zid}
        AND gldetail.zid = {zid}
        AND cacus.zid = {zid}
        AND gldetail.xproj = 'GULSHAN TRADING'
        AND gldetail.xvoucher NOT LIKE '%%OB%%'
        AND glheader.xdate <= '{today_date}'
        AND gldetail.xsub IN {get_customer_id}
    GROUP BY gldetail.xsub
    
"""


# Execute the query using psycopg2 not pd.read sql
df_prev_balance_customer = get_data (balance_query , fetchOneOrAll = 'all' , df = True, conn = "local")
df_prev_balance_customer

# %%
# now merge with df_send_to_salesman
df_send_to_salesman_with_balance = pd.merge(df_send_to_salesman, df_prev_balance_customer, on = 'xcus', how='left')
df_send_to_salesman_with_balance

# # for send to salesman
# df_send_to_salesman = df_send_to_salesman.drop(columns=['xcus'])
df_send_to_salesman_with_balance
# filter out which customer balance is less than 100
df_send_to_salesman_with_balance= df_send_to_salesman_with_balance[df_send_to_salesman_with_balance['balance_till_today'] > 300]
df_send_to_salesman_with_balance

# %%

# get last payment date
last_payment_query = f"""SELECT glheader.xdate as last_payment_date,gldetail.xsub AS xcus, sum(gldetail.xprime) as last_payment
                        FROM glheader
                        JOIN gldetail ON glheader.xvoucher = gldetail.xvoucher
                        JOIN cacus ON gldetail.xsub = cacus.xcus
                        WHERE glheader.zid = {zid}
                            AND gldetail.zid = {zid}
                            AND cacus.zid = {zid}
                            AND gldetail.xproj = 'GULSHAN TRADING'
                            AND gldetail.xvoucher LIKE '%%RCT-%%'
                            AND glheader.xdate <= '{today_date}'
                            AND gldetail.xsub in {get_customer_id}
                            group by  glheader.xdate,gldetail.xsub
                    order by gldetail.xsub, glheader.xdate desc
                    """
df_last_payment = get_data (last_payment_query , fetchOneOrAll = 'all' , df = True, conn = "local")
df_last_payment = df_last_payment.drop_duplicates(subset=['xcus'], keep= 'first' )

# %%
df_last_payment['last_payment'] = abs (df_last_payment['last_payment'])
df_last_payment

# %%
# create last payment date column
df_send_to_salesman_with_balance = pd.merge(df_send_to_salesman_with_balance, df_last_payment, on = 'xcus', how='left')
df_send_to_salesman_with_balance


# %%
df_send_to_salesman_with_balance = df_send_to_salesman_with_balance.sort_values(by = ['xsp','xcus','date_diff'])
df_send_to_salesman_with_balance

# %%
# remove unnecessary column
df_send_to_salesman_with_balance = df_send_to_salesman_with_balance.iloc[:, [0,6,7,2,8,11,9,10,12,13,5]]
df_send_to_salesman_with_balance

# %%
df_send_to_salesman_with_balance = df_send_to_salesman_with_balance[df_send_to_salesman_with_balance['date_diff'] > pd.Timedelta(days=0)]\
                                    .sort_values(by= ['date_diff', 'balance_till_today'], ascending=False)

df_send_to_salesman_with_balance

# %%
df_send_to_salesman_with_balance.to_excel("H_71_2_reminder.xlsx", index=False)

# %%

# change color of worksheet base ond date difference condition
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the existing workbook
wb = load_workbook("H_71_2_reminder.xlsx")
ws = wb.active

# Define the column index for "date_diff"
date_diff_column_index = 8  # Column J

# Define conditions for coloring cells in the "date_diff" column
for row in range(2, ws.max_row + 1):
    date_diff_cell = ws.cell(row=row, column=date_diff_column_index)

    if date_diff_cell.value > 15:
        date_diff_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    elif 10 <= date_diff_cell.value <= 15:
        date_diff_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    else:
        date_diff_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green

# Save the modified workbook
wb.save("H_71_2_reminder.xlsx")


# %%


# %%
# send to salesman

for salesman_id, (email, salesman_name) in salesman_info_dict.items():
    print(salesman_id)

    # Filter the DataFrame for the specific salesman_id
    salesman_df = df_send_to_salesman_with_balance[df_send_to_salesman_with_balance['xsp'] == salesman_id]

    # Check if the DataFrame is not empty
    if not salesman_df.empty:
        # Create a unique Excel file for each salesman
        excel_file_path = f"salesman_{salesman_id}.xlsx"
        salesman_df.to_excel(excel_file_path, index=False)

        # Create the HTML body with the filtered DataFrame
        html_body = [(salesman_df, f" Customer Due balance from last 20 days to {today_date})")]

        # Send email with the Excel file as the attachment
        send_mail(
            subject=f"H_71.2 Customer due Balance details from last month till {today_date}",
            bodyText=f"""
            Dear {salesman_name} <br>
            Plese find herewith the attached file regarding customer due balance from last 20 dayas
             """,
            attachment=[excel_file_path],
            recipient=[email, ],
            html_body=html_body
        )
    else:
        print(f"No data for salesman: {salesman_id}")
# delete all excelfile from current directory


# %%
## %%
# For Riyad vai
excel_file_path = f"H_71_2_reminder.xlsx"

html_body = [(df_send_to_salesman_with_balance, f" Customer cumulative balance from last 20 days")]

send_mail(
    subject="H71.2 Customer cumulative balance to reminder salesman",
    bodyText=f"Todays {today_date} mail sent to all salesman and find the excel sheet of district delivered goods",
    attachment=[excel_file_path],
    recipient=['XXXXXX' ],
  
)


# %%
def delete_xlsx_files():
    try:
        for file in glob.glob("*.xlsx"):
            os.remove(file)
        print("All .xlsx files deleted successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Call the function to delete .xlsx files in the current working directory
delete_xlsx_files()

# %%



